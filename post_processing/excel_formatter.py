"""Module pour formater des données dans des fichiers Excel."""

import re
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any

# Importation des outils de style d'openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


class ExcelFormatter:
    """
    Formate et exporte des données structurées vers des fichiers Excel stylisés.
    """

    TEST_CASE_ID_PATTERN = re.compile(r"^CU\d{2}_SB\d{2}_C[PEL]\d{3}_.+(?<!_)$")

    # Définition des styles
    HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    CP_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert
    CE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge
    CL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Jaune

    def _validate_test_case_id(self, test_id: str) -> bool:
        """Valide le format de l'identifiant du cas de test."""
        return bool(self.TEST_CASE_ID_PATTERN.match(test_id))

    def format_test_matrix(
        self, test_cases: List[Dict[str, Any]], output_path: str
    ) -> List[str]:
        """
        Crée et formate un fichier Excel pour une matrice de tests.

        Args:
            test_cases: Une liste de dictionnaires, chaque dictionnaire représentant un cas de test.
                        Chaque dict doit contenir au moins 'id' et 'type'.
            output_path: Le chemin du fichier Excel à créer (ex: 'reports/matrice_tests.xlsx').

        Returns:
            Une liste d'erreurs de validation rencontrées.
        """
        errors = []
        for i, case in enumerate(test_cases):
            if not self._validate_test_case_id(case.get("id", "")):
                errors.append(f"Ligne {i+2}: L'ID du cas de test '{case.get('id')}' ne respecte pas le format requis.")

        # Créer un DataFrame pandas
        df = pd.DataFrame(test_cases)

        # S'assurer que le répertoire de sortie existe
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Matrice de Tests")
            worksheet = writer.sheets["Matrice de Tests"]

            # Appliquer les styles
            self._apply_styles(worksheet)

        return errors

    def _apply_styles(self, worksheet):
        """
        Applique le formatage conditionnel et les styles à la feuille de calcul.
        """
        # Style des en-têtes
        for cell in worksheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Formatage conditionnel des lignes
        type_col_idx = -1
        for idx, cell in enumerate(worksheet[1]):
            if cell.value == 'type':
                type_col_idx = idx + 1
                break
        
        if type_col_idx != -1:
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                cell_type = row[type_col_idx - 1].value
                fill_style = None
                if cell_type == "CP":
                    fill_style = self.CP_FILL
                elif cell_type == "CE":
                    fill_style = self.CE_FILL
                elif cell_type == "CL":
                    fill_style = self.CL_FILL
                
                if fill_style:
                    for cell in row:
                        cell.fill = fill_style

        # Ajustement de la largeur des colonnes
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except: # noqa
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width


# --- Démonstration --- #
async def main():
    formatter = ExcelFormatter()
    test_data = [
        {
            "id": "CU01_SB01_CP001_connexion_valide",
            "description": "Vérifier la connexion avec un utilisateur et un mot de passe valides.",
            "type": "CP"
        },
        {
            "id": "CU01_SB01_CE001_mot_de_passe_incorrect",
            "description": "Vérifier le message d'erreur avec un mot de passe incorrect.",
            "type": "CE"
        },
        {
            "id": "CU01_SB02_CL001_champ_email_vide",
            "description": "Vérifier la réaction du système quand le champ email est laissé vide.",
            "type": "CL"
        },
        {
            "id": "ID_INVALIDE",
            "description": "Ce cas a un ID incorrect et devrait être signalé.",
            "type": "CP"
        }
    ]

    output_file = "reports/matrice_de_test_formatee.xlsx"
    print(f"Génération du fichier Excel de démonstration : {output_file}")

    validation_errors = formatter.format_test_matrix(test_data, output_file)

    if validation_errors:
        print("\nErreurs de validation détectées :")
        for error in validation_errors:
            print(f"- {error}")
    else:
        print("\nAucune erreur de validation.")

    print("\nFichier Excel généré avec succès.")


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())